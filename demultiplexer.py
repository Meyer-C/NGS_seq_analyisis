import gzip
from Bio import SeqIO
from openpyxl import load_workbook
from zipfile import ZipFile
import time

start = time.time()

# open zipped fastq file
all_seqs = []
# just remember actual files will require gzip piece right here

with gzip.open('top_10K_reads_cleaned_2021i1.fq.gz','rt') as a:
    for record in SeqIO.parse(a,'fastq'):
        all_seqs.append(str(record.seq))
print('File has been unzipped\n')

# pull the trap tag sequences
# open the excel workbook
wb = load_workbook(('20210805_trap_tag_seqs.xlsx'))
# identify the sheet being used
seq_sheet = wb['Sheet1']
tag_info = []
trap_tags = []
# pulls all the information from an excel spreadsheet
for x in range(2,102):
    tag_info.append((seq_sheet.cell(row=x,column=1).value))
    trap_tags.append((seq_sheet.cell(row=x,column=2).value))
print('Trap tag sequences have been identified\n')

# identify the target sequence and upstream constant sequence
# in the demultiplexer we really just need the upstream constant so this part can be shortened
target_seq = 'CCAGTCCTCAACAAGCTGCGCCGGTACTAGGCGACTCGGCACAGACGGTCTTACGCGTGCGGGCTACGTCCGACTATACCTTTGGTGGTAACCGGCTTCAACNNNNNNNNGCGATGAGGAACCCAGAA'
us_const = target_seq[84:102]

def demultiplexing(all_seqs,trap_tags,us_const):
    # this just creates a list of lists with length being the number of possible trap tags
    # filled with sequences that match the trap tag. trap tag index is based on it's number identifier
    seq_collections = [[] for i in range(len(trap_tags))]
    tag_in_seq = []
    tag_indexes = []
    all_tag_indexes = []
    # pull out all the tags from possible trap tags
    for a in range(len(trap_tags)):
        tag = trap_tags[a]
        # pull out each individual sequence
        for x in range(10000):
            seq = all_seqs[x]
            # index through the sequence to find every possible subsequence
            for y in range(len(seq)-len(us_const)):
                subseq = seq[y:y+len(us_const)]
                # compare the subsequences to the upstream constant
                # if it's the same as the upstream constant then take the next 8 bases as the found trap tag sequence
                if subseq == us_const:
                    found_tag = (seq[(y+len(us_const)):(y+len(us_const)+8)])
                    # if the found trap tag sequence == one of the trap tags, append the whole sequence to the position
                    # of that tag in the list of lists
                    if found_tag == tag:
                        seq_collections[a].append(seq)
                        tag_in_seq.append(found_tag)
                        all_tag_indexes.append(a)
    # create indexes for all of the used trap tags and remove any repeated indexes
    # in final_subseq_search it will only run the subsequence search on the indexes that are filled with sequences to
    # make it run a bit faster
    for z in all_tag_indexes:
        if z not in tag_indexes:
            tag_indexes.append(z)

    return seq_collections,tag_indexes


demulti = demultiplexing(all_seqs,trap_tags,us_const)

# creating output information
output = str(demulti[0])
# trap tag numbers for readability in the output info file
tag_index = demulti[1]
tag_num = []
for x in range(len(tag_index)):
    tag_num.append(tag_index[x]+1)
output_info = str(demulti[1])

# this is a long way to write all of the outputs to a zip file so it doesn't take up too much space, couldn't find an
# efficient way to do this so this is the best I could come up with

# writing all output information to a .txt file to be zipped later
with open('demultiplexed_sequences.txt','w') as demulti_seqs:
    demulti_seqs.write(output_info + '\n' + output)

# moving .txt file to a zipped file to save space
zipped = ZipFile('demultiplexed_sequences.zip','w')
zipped.write('demultiplexed_sequences.txt')
zipped.close()

# erasing original .txt file to save space
with open('demultiplexed_sequences.txt','w') as demulti_seqs:
    demulti_seqs.write('Demultiplexed sequences have been moved to the zipped file "demultiplexed_sequences.zip"\n'
                       'The trap tags found were: ' + str(tag_num))

# find runtime and tell that the program has finished running
end = time.time()
print('Demultiplexed sequences have been moved to the zipped file "demultiplexed_sequences.zip"\n'
      'The trap tags found were: ' + str(tag_num) + '\nThe total runtime was: ' + str(end-start))







